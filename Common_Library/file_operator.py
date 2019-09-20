# -*- coding: utf-8 -*-
# @Time   : 2019/5/13 14:00
# @Author  : Bamboo.pan
# @Email   : Bamboo.pan@hp.com
# @File    : test.py
# @Project : demo
import sys
import os
import shutil
import yaml
from openpyxl import load_workbook


class File:
    def __init__(self, folder_path, name, size=0):
        self.folder_path = folder_path
        self.name = name
        self.current_file = os.path.join(self.folder_path, self.name)
        self.size = size

    def new(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def open(self):
        f = open(self.current_file)
        return f

    def read(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def write(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def close(self):
        print(sys._getframe().f_code.co_name + " " + self.name + "  finished")

    def copy(self, des):
        new_file = shutil.copyfile(self.current_file, des)
        return new_file

    def move(self, des):
        new_file = shutil.move(self.current_file, des)
        return new_file

    def delete(self):
        os.remove(self.current_file)
        if not self.exist():
            return True
        else:
            return False

    def rename(self, new_name):
        new_file = os.rename(self.current_file, os.path.join(self.folder_path, new_name))
        return new_file

    def exist(self):
        result = os.path.isfile(self.current_file)
        return result


class XlsxFile(File):
    def __init__(self,
                 folder_path,
                 name,
                 size=0,
                 sheet_name='Sheet1',
                 rows=1,
                 cols=1):
        File.__init__(self, folder_path, name, size)
        self.active_sheet = ''
        self.sheet_name = sheet_name
        self.rows = rows
        self.cols = cols

    def open(self):
        excel_handle = load_workbook(os.path.join(self.folder_path, self.name))
        return excel_handle

    def read(self, excel_handle):
        config_sheet = excel_handle['config']
        dic = {}
        for i in range(2, self.get_rows(config_sheet) + 1):
            dic[config_sheet.cell(row=i,
                                  column=1).value.lower()] = config_sheet.cell(
                row=i, column=2).value.lower()
        dic['email'] = dic['email'].split(';')
        if dic['needbuild'] == 'y':
            dic['needbuild'] = True
        else:
            dic['needbuild'] = False
        uut_sheet = excel_handle['uut']
        uut_list = []
        for i in range(2, self.get_rows(uut_sheet) + 1):
            uut_dic = {}
            for j in range(1, self.get_cols(uut_sheet) + 1):
                if uut_sheet.cell(row=i, column=j).value is None:
                    continue
                uut_dic[uut_sheet.cell(
                    row=1, column=j).value.lower()] = uut_sheet.cell(
                    row=i, column=j).value.lower()
            if len(uut_dic) != self.get_cols(uut_sheet):
                continue
            uut_list.append(uut_dic)
        dic['uutlist'] = uut_list
        scripts_sheet = excel_handle['scripts']
        scripts_list = []
        scripts_dic = {}
        for i in range(2, self.get_rows(scripts_sheet) + 1):
            scripts_dic[scripts_sheet.cell(
                row=i, column=1).value] = scripts_sheet.cell(row=i,
                                                             column=2).value
        for key in scripts_dic.keys():
            if scripts_dic[key] == 'Y':
                scripts_list.append(key)
        dic['testscripts'] = list(filter(None, scripts_list))
        return dic

    def close(self, sheet_handle):
        sheet_handle.save(os.path.join(self.folder_path, self.name))

    def get_sheet_name(self, excel_handle):
        return excel_handle.sheetnames

    def get_rows(self, sheet_handle):
        return sheet_handle.max_row

    def get_cols(self, sheet_handle):
        return sheet_handle.max_column

    # TODO Excel Sheet
    def is_sheet_exist(self, sheet_name):
        pass

    def get_sheet(self, sheet_name):
        pass

    def add_sheet(self, sheet_name):
        pass

    def del_sheet(self, sheet_name):
        pass

    def rename_sheet(self, old_name, new_name):
        pass

    # TODO Excel Cell
    def read_cell(self, x, y):
        pass

    def write_cell(self, x, y):
        pass


class YamlFile(File):

    def read(self, file_handle):
        res = yaml.safe_load(file_handle)
        return res

    def close(self, file_handle):
        file_handle.close()

    @staticmethod
    def save(data, save_path):
        with open(save_path, "w") as f:
            yaml.safe_dump(data, f)


class HtmlFile(File):
    pass


class TxtFile(File):
    pass


if __name__ == "__main__":
    folder_path = r"C:\Users\sich\Desktop\test_folder\1"
    tar = r"C:\Users\sich\Desktop\test_folder\2\1233.xlsx"
    name = "test.xlsx"

    f = File(folder_path, name)
    newf = f.rename('test2.xlsx')
    print(newf)
