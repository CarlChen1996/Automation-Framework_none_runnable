# -*- coding: utf-8 -*-
# @Time   : 2019/5/13 14:00
# @Author  : Bamboo.pan
# @Email   : Bamboo.pan@hp.com
# @File    : test.py
# @Project : demo
import sys
import os
import yaml
from openpyxl import load_workbook


class File:
    def __init__(self, folder_path, name, size=0):
        self.folder_path = folder_path
        self.name = name
        self.size = size

    def open(self):
        file_name = os.path.join(self.folder_path, self.name)
        f = open(file_name)
        # print("open {} pass".format(file_name))
        return f

    def read(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def write(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def close(self):

        print(sys._getframe().f_code.co_name + " " + self.name + "  finished")

    def copy(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def move(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def delete(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def rename(self):
        print(sys._getframe().f_code.co_name + "  finished")

    def exist(self):
        print(sys._getframe().f_code.co_name + "  finished")


class XlsxFile(File):
    def __init__(self, folder_path, name, size=0, sheet_name='Sheet1', rows=1, cols=1):
        File.__init__(self, folder_path, name, size)
        self.sheet_name = sheet_name
        self.rows = rows
        self.cols = cols

    def open(self):
        excel_handle = load_workbook(os.path.join(self.folder_path, self.name))
        return excel_handle

    def read(self, excel_handle):
        config_sheet = excel_handle['config']
        dic = {}
        for i in range(1, self.get_rows(config_sheet)+1):
            dic[config_sheet.cell(row=i, column=1).value] = config_sheet.cell(row=i, column=2).value
        dic['email'] = dic['email'].split()
        if dic['needbuild'] == 'Y':
            dic['needbuild'] = True
        else:
            dic['needbuild'] = False
        uutlist_sheet = excel_handle['uutlist']
        uut_list = []
        uut_sheet_rows = self.get_rows(uutlist_sheet)
        uut_sheet_cols = self.get_cols(uutlist_sheet)
        for i in range(2, uut_sheet_rows+1):
            uut_dic = {}
            for j in range(1, uut_sheet_cols+1):
                uut_dic[uutlist_sheet.cell(row=1, column=j).value] = uutlist_sheet.cell(row=i, column=j).value
            uut_list.append(uut_dic)
        dic['uutlist'] = uut_list
        scriptslist_sheet = excel_handle['scriptslist']
        scripts_list = []
        scripts_dic = {}
        scripts_sheet_rows = self.get_rows(scriptslist_sheet)
        for i in range(2, scripts_sheet_rows+1):
            scripts_dic[scriptslist_sheet.cell(row=i, column=1).value] = scriptslist_sheet.cell(row=i, column=2).value
        for key in scripts_dic.keys():
            if scripts_dic[key] == 'Y':
                scripts_list.append(key)
        dic['testscripts'] = scripts_list
        return dic


    def close(self, sheet_handle):
        sheet_handle.save(os.path.join(self.folder_path, self.name))

    def get_sheet_name(self, excel_handle):
        return excel_handle.sheetnames

    def get_rows(self, sheet_handle):
        return sheet_handle.max_row

    def get_cols(self, sheet_handle):
        return sheet_handle.max_column


class MsgFile(File):
    def __init__(self, folder_path, name, size, subject, receiver, sender,
                 send_date, content, attachment):
        File.__init__(self, folder_path, name, size)
        self.subject = subject
        self.receiver = receiver
        self.sender = sender
        self.send_date = send_date
        self.content = content
        self.attachment = attachment

    def get_attanchment(self):
        print(sys._getframe().f_code.co_name + "  finished")


class YamlFile(File):
    def read(self, file_handle):
        res = yaml.safe_load(file_handle)
        return res

    def close(self, file_handle):
        file_handle.close()


class HtmlFile(File):
    pass


class TxtFile(File):
    pass


if __name__ == "__main__":

    folder_path = "c:\\test"
    name = "123.txt"
    size = 0

    # f=File(folder_path,name,size)
    # f.open()
    # f.read()
    # f.close()

    # sheetname="a"
    # colum=1
    # row=2
    # xx=XLSX(folder_path,name,size,sheetname,row,colum)
    # xx.open()
    # xx._()
    # xx.getRows()
    # xx.close()

    # subject="bamboo's mail"
    # receiver="bamboo.pan@hp.com"
    # sender="bamboo1@hp.com"
    # senddate="20190501"
    # content="test mail"
    # attachment="123.txt"
    #
    # msg=MSG(folder_path,name,size,subject,receiver,sender,senddate,content,attachment)
    # msg.open()
    # msg.getAttanchment()
    # msg.close()

    txt = TxtFile(folder_path, name, size)
    txt.open()
    txt.read()
    txt.close()
