# -*- coding: utf-8 -*-
# @Time   : 2019/7/15 10:32
# @Author  : balance.cheng
# @Email   : balance.cheng@hp.com
# @File    : QTPutils.py
# @Project : demo
import time
import win32com.client
import os
import openpyxl
# from Framework_Kernel.analyzer import Analyzer
from Framework_Kernel.analyzer import FrameworkSettings
from Common_Library.file_transfer import FTPUtils


class HPDMOperator:
    def __init__(self):
        self.__load_config()
        self.__ip = ''

    def __load_config(self):
        # -----------FTP settings ----------------
        config_ftp = FrameworkSettings().ftp_settings
        self.__ftp = config_ftp['server_address']
        self.__ftp_user = config_ftp['username']
        self.__ftp_passwd = config_ftp['password']
        # ----------QTP settings------------------
        config_qtp = FrameworkSettings().qtp_settings
        config_qtp_script = config_qtp['scripts_path']
        self.__test_data_path = r'Configuration\{}'.format(config_qtp['test_data'])
        self.__create_filter_path = config_qtp_script['create_filter']
        self.__send_command_path = config_qtp_script['send_command']
        self.__send_packages_path = config_qtp_script['send_packages']
        self.__discover_devices_path = config_qtp_script['discover_devices']
        self.__get_result_path = config_qtp_script['get_result']
        # ---------HPDM settings ----------------------
        self.__os_list = FrameworkSettings().hpdm_settings['os_list']
        self.__repository_path = FrameworkSettings().hpdm_settings['repository_path']

    # General Excel file as QTP DataTable
    def __initial_test_data(self, task):
        workbook = openpyxl.load_workbook(self.__test_data_path)
        empty_workbook = self.__clear_data(workbook)
        if empty_workbook:
            final_workbook = self.__fill_data(empty_workbook, task)
            final_workbook.close()
        return True

    def __clear_data(self, data_workbook):
        for os_item in self.__os_list.values():
            sheet_uut = data_workbook['UUT_{}'.format(os_item)]
            rows = sheet_uut.max_row
            for i in range(2, rows + 1):
                """
                Delete all the data in sheet
                """
                sheet_uut.delete_rows(i)
                data_workbook.save(self.__test_data_path)
        return data_workbook

    def __fill_data(self, data_workbook, task):
        # ----------Set UUT Data according task uut_list---------------------------------
        for uut in task.get_uut_list():
            if uut.get_version().upper() in self.__os_list.keys():
                sheet_uut = data_workbook['UUT_{}'.format(self.__os_list[uut.get_version().upper()])]
                max_row = sheet_uut.max_row
                # print(uut.get_version(),sheet_uut.max_row)
                sheet_uut.cell(max_row + 1, 1).value = uut.get_ip()
                sheet_uut.cell(max_row + 1, 2).value = uut.get_mac().upper()
                print('===========================================')
                print(uut.get_ip(), uut.get_mac())
                print('===========================================')
                data_workbook.save(self.__test_data_path)
        # ---------Set Config Data according task uut_list exe_list ------------------
        package_path = os.path.dirname(task.get_exe_file_list()[0])  # current only one exe file for one task
        sheet_config = data_workbook['Config']
        if 'LINUX' in package_path.upper():
            sheet_config.cell(2, 1).value = "Linux_{}".format(task.get_name())
        elif 'WINDOW' in package_path.upper():
            sheet_config.cell(2, 1).value = "Windows_{}".format(task.get_name())
        else:
            sheet_config.cell(2, 1).value = "Unknown_{}".format(task.get_name())
        # package_path given by build start with "/", so below path has no "/"
        sheet_config.cell(2, 2).value = '{}{}'.format(self.__repository_path, package_path)
        data_workbook.save(self.__test_data_path)
        return data_workbook

    def __upload_test_data(self, deploy_host):
        ftp = FTPUtils(self.__ftp, self.__ftp_user, self.__ftp_passwd)
        print(ftp.get_item_list(''))
        ftp.upload_file(file_name=self.__test_data_path, save_as_name='{}.xlsx'.format(deploy_host.get_hostname()))
        ftp.close()

    def __run_qtp_script(self, test_path):
        import pythoncom
        pythoncom.CoInitialize()
        qtp = win32com.client.DispatchEx("QuickTest.Application", self.__ip)
        qtp.Launch()
        qtp.Visible = True
        qtp.Open(test_path)
        qtp.Test.Run()
        qtp.Quit()
        pythoncom.CoUninitialize()
        time.sleep(5)

    def __create_filter(self):
        self.__run_qtp_script(self.__create_filter_path)

    def discover_devices(self, task):
        self.__run_qtp_script(self.__discover_devices_path)

    def deploy_task(self, task, deploy_host):
        self.__ip = deploy_host.get_ip()
        if self.__initial_test_data(task):
            self.__upload_test_data(deploy_host)
        self.discover_devices(task)
        self.__run_qtp_script(self.__send_packages_path)

    def execute_task(self, host):
        self.__ip = host.get_ip()
        self.__run_qtp_script(self.__send_command_path)

    def get_result(self, host):
        self.__ip = host.get_ip()
        self.__run_qtp_script(self.__get_result_path)
